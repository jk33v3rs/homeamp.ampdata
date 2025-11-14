"""
Pulumi-Based Plugin Update Monitor

Automated hourly plugin update detection system that:
1. Scans plugin repositories (GitHub, SpigotMC, Bukkit)
2. Detects new plugin versions
3. Stages updates (does NOT automatically deploy)
4. Writes results to Excel spreadsheet
5. Integrates with CI/CD systems

This is STAGING ONLY - updates are detected and prepared but require
explicit admin approval before deployment.
"""

from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
from datetime import datetime, timedelta
import logging
import json
import asyncio
import aiohttp
from dataclasses import dataclass, asdict


@dataclass
class PluginUpdate:
    """Plugin update information"""
    plugin_name: str
    current_version: str
    latest_version: str
    download_url: str
    release_date: str
    changelog: str
    repository_type: str  # 'github', 'spigot', 'bukkit', 'jenkins'
    repository_id: str
    file_size_bytes: int
    requires_restart: bool
    breaking_changes: bool
    compatibility_notes: str


@dataclass
class StagingEntry:
    """Staged update entry"""
    update_id: str
    plugin_update: PluginUpdate
    staged_at: str
    staged_by: str  # 'automated' or admin username
    approval_status: str  # 'pending', 'approved', 'rejected'
    approved_by: Optional[str]
    approved_at: Optional[str]
    deployment_status: str  # 'staged', 'deploying', 'deployed', 'failed'
    notes: str


class PluginUpdateMonitor:
    """
    Monitors plugin repositories for updates and stages them for review.
    Integrates with Pulumi for infrastructure automation.
    """
    
    def __init__(self, 
                 staging_path: Path,
                 plugin_registry_path: Path,
                 excel_output_path: Optional[Path] = None):
        """
        Initialize update monitor
        
        Args:
            staging_path: Directory for staged plugin JARs
            plugin_registry_path: Path to plugin_registry.yaml
            excel_output_path: Path to Excel spreadsheet for results
        """
        self.staging_path = staging_path
        self.plugin_registry_path = plugin_registry_path
        self.excel_output_path = excel_output_path
        self.logger = logging.getLogger(__name__)
        
        # Create staging directories
        self.staging_path.mkdir(parents=True, exist_ok=True)
        self.staging_metadata_path = self.staging_path / "metadata"
        self.staging_metadata_path.mkdir(exist_ok=True)
        
        # Load plugin registry
        self.plugin_registry = self._load_plugin_registry()
        
        # HTTP session for API calls
        self.session: Optional[aiohttp.ClientSession] = None
    
    def _load_plugin_registry(self) -> Dict[str, Dict[str, Any]]:
        """Load plugin registry with repository endpoints"""
        import yaml
        
        if not self.plugin_registry_path.exists():
            self.logger.warning(f"Plugin registry not found: {self.plugin_registry_path}")
            return {}
        
        try:
            with open(self.plugin_registry_path, 'r') as f:
                registry = yaml.safe_load(f) or {}
            self.logger.info(f"Loaded {len(registry)} plugins from registry")
            return registry
        except Exception as e:
            self.logger.error(f"Failed to load plugin registry: {e}")
            return {}
    
    async def check_all_updates(self) -> List[PluginUpdate]:
        """
        Check all registered plugins for updates
        
        Returns:
            List of available plugin updates
        """
        self.logger.info("Starting plugin update check...")
        updates = []
        
        async with aiohttp.ClientSession() as self.session:
            tasks = []
            for plugin_name, plugin_info in self.plugin_registry.items():
                task = self._check_plugin_update(plugin_name, plugin_info)
                tasks.append(task)
            
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            for result in results:
                if isinstance(result, Exception):
                    self.logger.error(f"Update check failed: {result}")
                elif result is not None:
                    updates.append(result)
        
        self.logger.info(f"Found {len(updates)} plugin updates")
        return updates
    
    async def _check_plugin_update(self, 
                                   plugin_name: str, 
                                   plugin_info: Dict[str, Any]) -> Optional[PluginUpdate]:
        """Check single plugin for updates"""
        try:
            repo_type = plugin_info.get('repository_type')
            current_version = plugin_info.get('current_version', '0.0.0')
            
            if repo_type == 'github':
                return await self._check_github_release(plugin_name, plugin_info, current_version)
            elif repo_type == 'spigot':
                return await self._check_spigot_resource(plugin_name, plugin_info, current_version)
            elif repo_type == 'bukkit':
                return await self._check_bukkit_resource(plugin_name, plugin_info, current_version)
            elif repo_type == 'jenkins':
                return await self._check_jenkins_build(plugin_name, plugin_info, current_version)
            else:
                self.logger.warning(f"Unknown repository type for {plugin_name}: {repo_type}")
                return None
        
        except Exception as e:
            self.logger.error(f"Failed to check {plugin_name}: {e}")
            return None
    
    async def _check_github_release(self, 
                                    plugin_name: str, 
                                    plugin_info: Dict[str, Any],
                                    current_version: str) -> Optional[PluginUpdate]:
        """Check GitHub releases for updates"""
        repo = plugin_info.get('github_repo')
        if not repo:
            return None
        
        url = f"https://api.github.com/repos/{repo}/releases/latest"
        headers = {"Accept": "application/vnd.github+json"}
        
        try:
            async with self.session.get(url, headers=headers, timeout=30) as response:
                if response.status != 200:
                    return None
                
                data = await response.json()
                latest_version = data.get('tag_name', '').lstrip('v')
                
                # Check if update available
                if not self._is_newer_version(current_version, latest_version):
                    return None
                
                # Find JAR asset
                download_url = None
                file_size = 0
                for asset in data.get('assets', []):
                    if asset['name'].endswith('.jar'):
                        download_url = asset['browser_download_url']
                        file_size = asset.get('size', 0)
                        break
                
                if not download_url:
                    self.logger.warning(f"No JAR found in {plugin_name} release")
                    return None
                
                return PluginUpdate(
                    plugin_name=plugin_name,
                    current_version=current_version,
                    latest_version=latest_version,
                    download_url=download_url,
                    release_date=data.get('published_at', ''),
                    changelog=data.get('body', ''),
                    repository_type='github',
                    repository_id=repo,
                    file_size_bytes=file_size,
                    requires_restart=plugin_info.get('requires_restart', True),
                    breaking_changes=self._detect_breaking_changes(data.get('body', '')),
                    compatibility_notes=plugin_info.get('compatibility_notes', '')
                )
        
        except Exception as e:
            self.logger.error(f"GitHub check failed for {plugin_name}: {e}")
            return None
    
    async def _check_spigot_resource(self, 
                                     plugin_name: str, 
                                     plugin_info: Dict[str, Any],
                                     current_version: str) -> Optional[PluginUpdate]:
        """Check SpigotMC resource for updates"""
        resource_id = plugin_info.get('spigot_id')
        if not resource_id:
            return None
        
        # Spigot API endpoint
        url = f"https://api.spiget.org/v2/resources/{resource_id}/versions/latest"
        
        try:
            async with self.session.get(url, timeout=30) as response:
                if response.status != 200:
                    return None
                
                data = await response.json()
                latest_version = data.get('name', '')
                
                if not self._is_newer_version(current_version, latest_version):
                    return None
                
                # Get resource details
                resource_url = f"https://api.spiget.org/v2/resources/{resource_id}"
                async with self.session.get(resource_url, timeout=30) as res_response:
                    if res_response.status != 200:
                        return None
                    
                    resource_data = await res_response.json()
                    
                    return PluginUpdate(
                        plugin_name=plugin_name,
                        current_version=current_version,
                        latest_version=latest_version,
                        download_url=f"https://www.spigotmc.org/resources/{resource_id}/download?version={data.get('id')}",
                        release_date=datetime.fromtimestamp(data.get('releaseDate', 0)).isoformat(),
                        changelog=resource_data.get('updateMessage', ''),
                        repository_type='spigot',
                        repository_id=resource_id,
                        file_size_bytes=resource_data.get('file', {}).get('size', 0),
                        requires_restart=plugin_info.get('requires_restart', True),
                        breaking_changes=False,  # Spigot doesn't provide this info
                        compatibility_notes=plugin_info.get('compatibility_notes', '')
                    )
        
        except Exception as e:
            self.logger.error(f"Spigot check failed for {plugin_name}: {e}")
            return None
    
    async def _check_bukkit_resource(self, 
                                     plugin_name: str, 
                                     plugin_info: Dict[str, Any],
                                     current_version: str) -> Optional[PluginUpdate]:
        """Check Bukkit/CurseForge resource for updates"""
        project_id = plugin_info.get('bukkit_id')
        if not project_id:
            return None
        
        # Note: CurseForge API requires API key
        # This is a placeholder - implement when API key is available
        self.logger.warning(f"Bukkit check not implemented for {plugin_name}")
        return None
    
    async def _check_jenkins_build(self, 
                                   plugin_name: str, 
                                   plugin_info: Dict[str, Any],
                                   current_version: str) -> Optional[PluginUpdate]:
        """Check Jenkins CI for latest build"""
        jenkins_url = plugin_info.get('jenkins_url')
        if not jenkins_url:
            return None
        
        url = f"{jenkins_url}/lastSuccessfulBuild/api/json"
        
        try:
            async with self.session.get(url, timeout=30) as response:
                if response.status != 200:
                    return None
                
                data = await response.json()
                build_number = str(data.get('number', ''))
                
                if not self._is_newer_version(current_version, build_number):
                    return None
                
                # Find JAR artifact
                download_url = None
                file_size = 0
                for artifact in data.get('artifacts', []):
                    if artifact['fileName'].endswith('.jar'):
                        download_url = f"{jenkins_url}/lastSuccessfulBuild/artifact/{artifact['relativePath']}"
                        file_size = artifact.get('size', 0)
                        break
                
                if not download_url:
                    return None
                
                return PluginUpdate(
                    plugin_name=plugin_name,
                    current_version=current_version,
                    latest_version=f"build-{build_number}",
                    download_url=download_url,
                    release_date=datetime.fromtimestamp(data.get('timestamp', 0) / 1000).isoformat(),
                    changelog=data.get('changeSet', {}).get('comment', ''),
                    repository_type='jenkins',
                    repository_id=jenkins_url,
                    file_size_bytes=file_size,
                    requires_restart=plugin_info.get('requires_restart', True),
                    breaking_changes=False,
                    compatibility_notes=plugin_info.get('compatibility_notes', '')
                )
        
        except Exception as e:
            self.logger.error(f"Jenkins check failed for {plugin_name}: {e}")
            return None
    
    def _is_newer_version(self, current: str, latest: str) -> bool:
        """
        Compare version strings to determine if latest is newer
        
        Handles: semantic versioning (1.2.3), build numbers (build-123),
        and various version formats
        """
        try:
            # Remove common prefixes
            current = current.lstrip('vV ')
            latest = latest.lstrip('vV ')
            
            # Handle build numbers
            if current.startswith('build-') and latest.startswith('build-'):
                current_num = int(current.split('-')[1])
                latest_num = int(latest.split('-')[1])
                return latest_num > current_num
            
            # Handle semantic versioning
            def parse_version(v: str) -> Tuple[int, ...]:
                # Remove non-numeric suffixes (e.g., "1.2.3-SNAPSHOT")
                v = v.split('-')[0].split('+')[0]
                parts = v.split('.')
                return tuple(int(p) if p.isdigit() else 0 for p in parts)
            
            current_parts = parse_version(current)
            latest_parts = parse_version(latest)
            
            return latest_parts > current_parts
        
        except Exception as e:
            self.logger.error(f"Version comparison failed: {current} vs {latest}: {e}")
            # If can't determine, assume it's newer to be safe
            return True
    
    def _detect_breaking_changes(self, changelog: str) -> bool:
        """Detect if changelog mentions breaking changes"""
        breaking_keywords = [
            'breaking change',
            'breaking',
            'incompatible',
            'migration required',
            'not backward compatible'
        ]
        
        changelog_lower = changelog.lower()
        return any(keyword in changelog_lower for keyword in breaking_keywords)
    
    async def stage_update(self, update: PluginUpdate, staged_by: str = 'automated') -> StagingEntry:
        """
        Stage a plugin update for review
        
        Args:
            update: Plugin update to stage
            staged_by: Username who staged (default: 'automated')
            
        Returns:
            Staging entry with metadata
        """
        update_id = f"{update.plugin_name}-{update.latest_version}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        
        try:
            # Download plugin JAR to staging area
            download_path = self.staging_path / f"{update.plugin_name}-{update.latest_version}.jar"
            
            async with self.session.get(update.download_url, timeout=300) as response:
                if response.status != 200:
                    raise Exception(f"Download failed: HTTP {response.status}")
                
                with open(download_path, 'wb') as f:
                    f.write(await response.read())
            
            self.logger.info(f"Downloaded {update.plugin_name} {update.latest_version} to staging")
            
            # Create staging entry
            entry = StagingEntry(
                update_id=update_id,
                plugin_update=update,
                staged_at=datetime.now().isoformat(),
                staged_by=staged_by,
                approval_status='pending',
                approved_by=None,
                approved_at=None,
                deployment_status='staged',
                notes=''
            )
            
            # Save metadata
            metadata_path = self.staging_metadata_path / f"{update_id}.json"
            with open(metadata_path, 'w') as f:
                json.dump(asdict(entry), f, indent=2)
            
            return entry
        
        except Exception as e:
            self.logger.error(f"Failed to stage {update.plugin_name}: {e}")
            raise
    
    async def stage_all_updates(self, updates: List[PluginUpdate]) -> List[StagingEntry]:
        """Stage multiple updates"""
        entries = []
        
        async with aiohttp.ClientSession() as self.session:
            for update in updates:
                try:
                    entry = await self.stage_update(update)
                    entries.append(entry)
                except Exception as e:
                    self.logger.error(f"Failed to stage {update.plugin_name}: {e}")
        
        return entries
    
    def write_to_excel(self, updates: List[PluginUpdate], staged_entries: List[StagingEntry]):
        """
        Write update results to Excel spreadsheet
        
        Updates the plugin matrix with:
        - Latest available versions
        - Update availability status
        - Staging status
        - Approval status
        """
        if not self.excel_output_path:
            self.logger.warning("No Excel output path configured")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            
            # Load or create workbook
            if self.excel_output_path.exists():
                wb = openpyxl.load_workbook(self.excel_output_path)
            else:
                wb = openpyxl.Workbook()
            
            # Get or create "Plugin Updates" sheet
            if "Plugin Updates" in wb.sheetnames:
                ws = wb["Plugin Updates"]
            else:
                ws = wb.create_sheet("Plugin Updates")
            
            # Clear existing data
            ws.delete_rows(1, ws.max_row)
            
            # Write headers
            headers = [
                "Plugin Name", "Current Version", "Latest Version", 
                "Update Available", "Repository", "Release Date",
                "Staged", "Approval Status", "Deployment Status",
                "Breaking Changes", "Requires Restart", "Download URL"
            ]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Write update data
            staged_dict = {e.plugin_update.plugin_name: e for e in staged_entries}
            
            for update in updates:
                staged_entry = staged_dict.get(update.plugin_name)
                
                row = [
                    update.plugin_name,
                    update.current_version,
                    update.latest_version,
                    "YES",
                    update.repository_type,
                    update.release_date,
                    "YES" if staged_entry else "NO",
                    staged_entry.approval_status if staged_entry else "N/A",
                    staged_entry.deployment_status if staged_entry else "N/A",
                    "YES" if update.breaking_changes else "NO",
                    "YES" if update.requires_restart else "NO",
                    update.download_url
                ]
                ws.append(row)
                
                # Highlight breaking changes
                if update.breaking_changes:
                    last_row = ws.max_row
                    for cell in ws[last_row]:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(self.excel_output_path)
            self.logger.info(f"Wrote {len(updates)} updates to Excel: {self.excel_output_path}")
        
        except Exception as e:
            self.logger.error(f"Failed to write to Excel: {e}")
    
    async def run_hourly_check(self):
        """
        Run hourly update check cycle
        
        This is the main entry point for Pulumi automation:
        1. Check all plugins for updates
        2. Stage any new updates
        3. Write results to Excel
        """
        self.logger.info("=== Starting hourly plugin update check ===")
        
        try:
            # Check for updates
            updates = await self.check_all_updates()
            
            if not updates:
                self.logger.info("No plugin updates available")
                return
            
            # Stage updates
            self.logger.info(f"Staging {len(updates)} plugin updates...")
            staged_entries = await self.stage_all_updates(updates)
            
            # Write to Excel
            self.write_to_excel(updates, staged_entries)
            
            self.logger.info(f"=== Hourly check complete: {len(updates)} updates staged ===")
        
        except Exception as e:
            self.logger.error(f"Hourly check failed: {e}")
            raise


async def main():
    """Test/development entry point"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Example paths - adjust as needed
    staging_path = Path("/var/lib/archivesmp/plugin-staging")
    registry_path = Path("/etc/archivesmp/plugin_registry.yaml")
    excel_path = Path("/var/lib/archivesmp/reports/plugin_updates.xlsx")
    
    monitor = PluginUpdateMonitor(
        staging_path=staging_path,
        plugin_registry_path=registry_path,
        excel_output_path=excel_path
    )
    
    await monitor.run_hourly_check()


if __name__ == "__main__":
    asyncio.run(main())
